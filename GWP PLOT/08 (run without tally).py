import tkinter as tk
import tkinter.messagebox
from tkinter import ttk
from tkinter import filedialog
import customtkinter
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf
import openpyxl
from openpyxl import load_workbook, workbook
import warnings
import os
import PIL
from PIL import ImageTk, Image
import pandas as pd
import glob

###replace path later
###line 48
###line 227
###line 447
###line 357

customtkinter.set_appearance_mode("Dark")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"



#select excel
def read_excel():
    global total_gwp_floor, total_gwp_transfer, total_gwp_wall, total_gwp_col, total_gwp_ftg, total_gwp_MISC, total_gwp, total_slab_area_m2
    file_path = r'U:\Steve\LCA\temp xlsx\temp.xlsx'
    wb = load_workbook(filename=file_path)
    ws = wb.active
    # Read the cells from 'A5' to 'B11'
    data = []
    for row in ws.iter_rows(min_row=5, max_row=12, min_col=1, max_col=2):
        data.append([cell.value for cell in row])

    #x_value = (total_gwp_col, total_gwp_wall, total_gwp_transfer, total_gwp_floor, total_gwp_ftg, total_gwp_misc)
    # Print the data
    for row in data:
        if row[0] == 'Total Slab GWP':
            total_gwp_floor = float(row[1])
        elif row[0] == 'Total Transfer Slab GWP':
            total_gwp_transfer = float(row[1])
        elif row[0] == 'Total Wall GWP':
            total_gwp_wall= float(row[1])
        elif row[0] == 'Total COLUMN GWP':
            total_gwp_col= float(row[1])
        elif row[0] == 'Total FOUNDATION GWP':
            total_gwp_ftg= float(row[1])
        elif row[0] == 'Total MISC GWP':
            total_gwp_MISC= float(row[1])
        elif row[0] == 'Total GWP':
            total_gwp = float(row[1])
        elif row[0] == 'Total Slab Area M2':
            total_slab_area_m2 = float(row[1])



    print(total_gwp_floor, total_gwp_transfer, total_gwp_wall, total_gwp_col, total_gwp_ftg, total_gwp_MISC, total_gwp, total_slab_area_m2)
    tkinter.messagebox.showinfo("showinfo", "excel loaded successfully")
    return total_gwp_floor, total_gwp_transfer, total_gwp_wall, total_gwp_col, total_gwp_ftg, total_gwp_MISC, total_gwp, total_slab_area_m2



#buttons
def button_of_piechar():
    global total_gwp_per_area

    total_gwp_per_area = total_gwp



    #total_gwp_basementwall = round((sum(list_gwp_basementwall)) / total_slab_area_m2, 1)
    #total_gwp_shearwall = round((sum(list_gwp_shearwall)) / total_slab_area_m2, 1)

    x_value = (total_gwp_col, total_gwp_wall, total_gwp_transfer, total_gwp_floor, total_gwp_ftg, total_gwp_MISC)

    labels = ["COLUMN", "WALL", "SLAB", "TRANSFER", "FTG", "MISC"]
    mycolors = ['#3c78d8', '#6aa84f', '#f1c232', '#ff8040', '#999999', '#FF0000']
    explodes = (0, 0, 0.1, 0, 0, 0)

    plt.pie([total_gwp_col, total_gwp_wall, total_gwp_floor, total_gwp_transfer,total_gwp_ftg, total_gwp_MISC], explode=explodes,
            labels=labels, colors=mycolors, autopct='%.2f %%',)

    plt.savefig(r'U:\Steve\LCA\temp png\pie chart.png')

    plt.show()


    return x_value


def benchmark_color_residential(y_value):
    if type(y_value) is list:
        y_value = sum(y_value)
    else:
        pass
    if y_value < 280:
       mycolor = "#4a9232"
    elif 280 < y_value < 360:
        mycolor = "#8acf30"
    elif 360 < y_value < 440:
        mycolor = "#b9ec5b"
    elif 440 < y_value < 520:
        mycolor = "#ecf10e"
    elif 520 < y_value < 600:
        mycolor = "#ffae88"
    elif 600 < y_value < 680:
        mycolor = "#ff8040"
    else:
        mycolor = "red"

    return mycolor


## part of barchart
def values_list_add_zero(lista):
    lista.insert(0, 0)

def rate_list_add_string(lista):
    lista.insert(0, '')

def pass_data_to_barchart():
    list_data = []
    list_data_temp =[]
    path = r'U:\Steve\LCA\temp xlsx\temp.xlsx'
    wb = openpyxl.load_workbook(path)
    ws = wb['Sheet1']

    list_label = ['SD/CD', 'BUILDING PERMIT', 'TENDER', 'IFC']

    for cell in ws['B']:
        data = cell.value
        if data is None:
            list_data_temp.append('')
            pass
        elif data == 'None':
            list_data_temp.append('')
            pass
        else:
            list_data.append(float(data))
            list_data_temp.append(float(data))

    res = dict(zip(list_label, list_data_temp))


    return res

def button_of_barchart():
    global buildingstage
    labels = ['IFC', 'TENDER', 'BUILDING\nPERMIT', 'SCHEMATIC\nDESIGN']
    #values = [311.1, 389]
    # colors = ['black', 'black', 'black', 'black']
    list_d = pass_data_to_barchart()

    a = 0
    list_data = []
    list_data_final = []
    for k, v in list_d.items():
        buildingstage = k
        gwp = v
        if v == '':
            list_data.append(0)
            a = a + 1

        else:
            list_data_final.append(buildingstage)
            list_data.append(gwp)
            a = a + 1

    buildingstage = list_data_final[-1]

    if len(list_data) > 0:
        values = list_data
    else:
        values = [0, 0, 0, 0]

    #print(buildingstage)

    color_total = {'A++': '#3E7A29', 'A+': '#4a9232', 'A': '#8acf30', 'B': '#d6ef7c', 'C': '#ecf10e', 'D': '#ffae88',
                   'E': '#ff8040', 'F': '#ff0000', 'G': '#d20000', 'White': '#ffffff'}

    colors = []
    rate_total = []

    for value in values:
        if 50 >= value > 0:
            value_color = color_total['A++']
            colors.append(value_color)
            rate_total.append('A++')


        elif 100 >= value > 50:
            value_color = color_total['A+']
            colors.append(value_color)
            rate_total.append('A+')

        elif 150 >= value > 100:
            value_color = color_total['A']
            colors.append(value_color)
            rate_total.append('A')


        elif 200 >= value > 150:
            value_color = color_total['B']
            colors.append(value_color)
            rate_total.append('B')


        elif 250 >= value > 200:
            value_color = color_total['C']
            colors.append(value_color)
            rate_total.append('C')


        elif 300 >= value > 250:
            value_color = color_total['D']
            colors.append(value_color)
            rate_total.append('D')


        elif 350 >= value > 300:
            value_color = color_total['E']
            colors.append(value_color)
            rate_total.append('E')


        elif 400 >= value > 350:
            value_color = color_total['F']
            colors.append(value_color)
            rate_total.append('F')


        elif value >= 400:
            value_color = color_total['G']
            colors.append(value_color)
            rate_total.append('G')

        else:
            value_color = color_total['White']
            colors.append(value_color)
            rate_total.append('')

    # colors = ['#8bcf30', '#ff8040', '#ff0000', '#d20000']
    # rate = ['', '', 'E', 'G']

    # Create a horizontal bar chart with custom hex colors

    values.reverse()
    rate_total.reverse()
    colors.reverse()


    fig, ax = plt.subplots()
    bars = ax.barh(labels, values, color=colors, height=0.3)

    # Set the x-axis label
    ax.set_xlabel('GWP (kgCO2e/m²)')

    # Set the maximum x value
    ax.set_xlim(right=550)

    # if len(colors) <= 2:
    #     colors.append('#ff0000')
    #     colors.append('#ff0000')
    #     rate_total.append('d')
    #     rate_total.append('d')
    #
    # print(colors)
    # print(rate_total)

    for i, bar in enumerate(bars):
        ax.text(bar.get_width() + 0.3, bar.get_y() + bar.get_height() / 2, rate_total[i], ha='left', va='center')

    plt.vlines(x=300, ymin=0, ymax=3.2, colors='green', ls=':', lw=2, label='vline_single - full height')

    # Show the chart

    plt.savefig(r'U:\Steve\LCA\temp png\bar chart.png')

    plt.show()



def button_of_export_data():
    lista = []
    #print(text_buildingtype.get())
    v = text_buildingtype.get()
    #["Residentail", "Education", "Health Care", "Industrial", "Office", "Other"]
    if v == "Residentail":
        p = r'U:\Steve\LCA\data collection\GWP data - Residentail.xlsx'
    elif v == "Education":
        p = r'U:\Steve\LCA\data collection\GWP data - Education.xlsx'
    elif v == "Health Care":
        p = r'U:\Steve\LCA\data collection\GWP data - Health Care.xlsx'
    elif v == "Industrial":
        p = r'U:\Steve\LCA\data collection\GWP data - Industrial.xlsx'
    elif v == "Office":
        p = r'U:\Steve\LCA\data collection\GWP data - Office.xlsx'
    else:
        p = r'U:\Steve\LCA\data collection\GWP data - Other.xlsx'
    wb_excel = load_workbook(p)
    ws_excel = wb_excel['Sheet1']
    jobnumber = text_jobnumber.get()
    #buildingstage = text_buildingtype.get()
    lista.append(str(jobnumber))
    lista.append(buildingstage)
    lista.append(total_slab_area_m2)
    lista.append(total_gwp_floor)
    lista.append(total_gwp_transfer)
    lista.append(total_gwp_wall)
    lista.append(total_gwp_col)
    lista.append(total_gwp_ftg)
    lista.append(total_gwp_MISC)
    lista.append(total_gwp_per_area)
    #print(lista)
    #print(jobnumber, buildingstage, total_slab_area_m2, total_gwp_floor, total_gwp_transfer, total_gwp_wall, total_gwp_col, total_gwp_misc,total_gwp_per_area)
    ws_excel.append(lista)
    wb_excel.save(p)



    tkinter.messagebox.showinfo("showinfo", "data exported")

    app.destroy()




app = customtkinter.CTk()
app.geometry("580x500")
app.title("Glotman Simpson.py")



title_frame = customtkinter.CTkFrame(master=app)
title_frame.place(relx=0.28, rely=0.13, anchor=tkinter.CENTER)

title_label = customtkinter.CTkLabel(master=title_frame, justify=tkinter.LEFT,text="GWP Plot", width=180, height=60, text_font=("Roboto Medium", -28))
title_label.pack(pady=10, padx=10)

frame_1 = customtkinter.CTkFrame(master=app, width=200, height=300)
frame_1.place(relx=0.28, rely=0.57, anchor=tkinter.CENTER)

logo_frame = customtkinter.CTkFrame(master=app, width=200, height=200)
logo_frame.place(relx=0.7, rely=0.25, anchor=tkinter.CENTER)

logo = ImageTk.PhotoImage(Image.open(r"U:\Steve\LCA\one pager program\03\Glotman-logo.png"))

logo_label = customtkinter.CTkLabel(master=logo_frame, image=logo)
logo_label.pack(pady=10, padx=10)


frame_2 = customtkinter.CTkFrame(master=app, width=200, height=185)
frame_2.place(relx=0.7, rely=0.69, anchor=tkinter.CENTER)

# label = customtkinter.CTkLabel(master=app, text="One Pager", width=180, height=60, text_font=("Roboto Medium", -28))
# label.place(relx=0.28, rely=0.15, anchor=tkinter.CENTER)

#
button_readdata = customtkinter.CTkButton(master=frame_1, text="Load Excel", command=read_excel, width=160, height=60,text_font=("Roboto Medium", -20) )
button_readdata.place(relx=0.50, rely=0.2, anchor=tkinter.CENTER)
#
button_piechart = customtkinter.CTkButton(master=frame_1, text="Create Pie Chart", command=button_of_piechar, width=160, height=60,text_font=("Roboto Medium", -20) )
button_piechart.place(relx=0.50, rely=0.53, anchor=tkinter.CENTER)
#
button_dotchart = customtkinter.CTkButton(master=frame_1, text="Create Bar Chart", command=button_of_barchart, width=160, height=60,text_font=("Roboto Medium", -20) )
button_dotchart.place(relx=0.50, rely=0.85, anchor=tkinter.CENTER)
#
#button_exportdata = customtkinter.CTkButton(master=frame_1, text="Export Data", command=button_of_dotchar, width=160, height=60,text_font=("Roboto Medium", -20) )
#button_exportdata.place(relx=0.50, rely=0.80, anchor=tkinter.CENTER)

#
label_jobnumber = customtkinter.CTkLabel(master=frame_2, text="Job number", text_font=("Roboto Medium", -10) )
label_jobnumber.place(relx=0.82, rely=0.15, anchor=tkinter.CENTER)

text_jobnumber = customtkinter.CTkEntry(master=frame_2, width=120, height=30,text_font=("Roboto Medium", -10),)
text_jobnumber.place(relx=0.35, rely=0.15, anchor=tkinter.CENTER)

label_buildingtype = customtkinter.CTkLabel(master=frame_2, text="building type", text_font=("Roboto Medium", -10) )
label_buildingtype.place(relx=0.82, rely=0.4, anchor=tkinter.CENTER)

buildingtype_selection = ["", "Residentail", "Education", "Health Care", "Industrial", "Office", "Other"]
text_buildingtype = customtkinter.CTkComboBox(values=buildingtype_selection, master=frame_2, width=120, height=30,text_font=("Roboto Medium", -10),)
text_buildingtype.place(relx=0.35, rely=0.4, anchor=tkinter.CENTER)

#
#button_buildingtype = customtkinter.CTkButton(master=frame_2, text="Building Type", command=button_of_dotchar, width=160, height=40,text_font=("Roboto Medium", -15) )
#button_buildingtype.place(relx=0.5, rely=0.7, anchor=tkinter.CENTER)


button_exportdata = customtkinter.CTkButton(master=frame_2, text="Export Data", command=button_of_export_data, width=160, height=60,text_font=("Roboto Medium", -20) )
button_exportdata.place(relx=0.50, rely=0.74, anchor=tkinter.CENTER)

app.mainloop()
