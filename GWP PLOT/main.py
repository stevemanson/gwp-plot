import tkinter as tk
import tkinter.messagebox
from tkinter import ttk
from tkinter import filedialog
import customtkinter
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf
import openpyxl
from openpyxl import load_workbook, workbook
import warnings
import os
import PIL
from PIL import ImageTk, Image
import pandas as pd
import glob




customtkinter.set_appearance_mode("Dark")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"



#select excel
def load_data():
    global tally_xlsx_path
    warnings.simplefilter(action='ignore', category=UserWarning)
    filename = filedialog.askopenfilename(title='Please select a excel report from Tally')
    path = filename
    tally_xlsx_path = path.replace('\\', '/')
    return tally_xlsx_path

def read_excel():
    global df1, df2, df3, df4, temp_csv_path
    root_not_used = tkinter.Tk()
    root_not_used.withdraw()
    #filename = filedialog.askopenfilename(title='Please select a excel report from Tally')
    #wb = load_data(filename)
    # print(filename)
    # print(wb)
    all_sheets = pd.read_excel(load_data(), sheet_name=None)
    sheets = all_sheets.keys()
    x = 0
    temp_csv_path = r'U:\Steve\LCA\temp csv'
    for sheet_name1 in sheets:
        x = x + 1
        sheet = pd.read_excel(tally_xlsx_path, sheet_name=sheet_name1)
        sheet.to_csv(temp_csv_path + '\\' + str(x) + "-" + sheet_name1 + ".csv", index=False)
        sheet_to_csv_path = (str(temp_csv_path + '\\' + str(x) + "-" + sheet_name1 + ".csv"))
        if "Report Summary" in sheet_to_csv_path:
            df1 = pd.read_csv(sheet_to_csv_path, header=0)
        if "Revit model" in sheet_to_csv_path:
            df2 = pd.read_csv(sheet_to_csv_path, header=1)
        if "Stage-Division" in sheet_to_csv_path:
            df3 = pd.read_csv(sheet_to_csv_path, header=0)
        if "Category-Family" in sheet_to_csv_path:
            df4 = pd.read_csv(sheet_to_csv_path, header=1)
    tkinter.messagebox.showinfo("showinfo", "excel loaded successfully")




    #print(df1.iat[7,1])
    #print(df2.iat[2, 3])
    # print(df2["Sum of Global Warming Potential Total (kgCO2eq)"].values[1])
    # print(df4['Row Labels'])


    return df1, df2, df3, df4



#buttons
def button_of_piechar():
    global total_slab_area_m2, total_gwp_col,total_gwp_wall,total_gwp_floor,total_gwp_ftg,total_gwp_misc,total_gwp_per_area, total_gwp_transfer
    #sheet1 = wb['Report Summary']
    #sheet2 = wb["Revit model"]
    #sheet4 = wb["Stage-Division"]
    #sheet10 = wb["Category-Family"]

    list_gwp_col = []
    list_gwp_wall = []
    list_gwp_floor = []
    list_gwp_ftg = []
    list_gwp_misc = []
    list_gwp_transfer = []
    list_gwp_basementwall = []
    list_gwp_shearwall = []
    list_gwp_parkingslab = []
    list_gwp_groundfloorslab = []
    list_gwp_towerslab = []
    list_gwp_ptslab = []


    slab_area = df1.iat[7,1]
    if "m²" in slab_area:
        slab_area_m2 = float(slab_area.replace('m²', ''))
        total_slab_area_m2 = round(slab_area_m2, 1)
    else:
        slab_area_m2 = float(slab_area.replace('ft²', ''))
        total_slab_area_m2 = round((slab_area_m2 * 0.092903), 1)

    total_gwp = float(df2["Sum of Global Warming Potential Total (kgCO2eq)"].values[1])
    total_gwp_per_area = round(total_gwp / total_slab_area_m2, 1)

    a = 0
    for string in df4['Row Labels']:
        if a >= 1:
            #print(a, "-", string)
            if "Slabband" in string:
                list_gwp_transfer.append(float(df4['Sum of Global Warming Potential Total (kgCO2eq)'].values[a]))
                a = a + 1
            elif "Column" in string:
                if "HSS" in string:
                    pass
                elif "Flange" in string:
                    pass
                else:
                    list_gwp_col.append(float(df4['Sum of Global Warming Potential Total (kgCO2eq)'].values[a]))
                a = a + 1
            elif "Shearwall" in string:
                list_gwp_wall.append(float(df4['Sum of Global Warming Potential Total (kgCO2eq)'].values[a]))
                a = a + 1
            elif "Other" in string:
                list_gwp_wall.append(float(df4['Sum of Global Warming Potential Total (kgCO2eq)'].values[a]))
                a = a + 1
            elif "Basement" in string:
                list_gwp_wall.append(float(df4['Sum of Global Warming Potential Total (kgCO2eq)'].values[a]))
                a = a + 1
            elif "Step" in string:
                list_gwp_wall.append(float(df4['Sum of Global Warming Potential Total (kgCO2eq)'].values[a]))
                a = a + 1
            elif "Zone" in string:
                list_gwp_wall.append(float(df4['Sum of Global Warming Potential Total (kgCO2eq)'].values[a]))
                a = a + 1
            elif "Upstand" in string:
                list_gwp_wall.append(float(df4['Sum of Global Warming Potential Total (kgCO2eq)'].values[a]))
                a = a + 1
            elif "Slab" in string:
                list_gwp_floor.append(float(df4['Sum of Global Warming Potential Total (kgCO2eq)'].values[a]))
                a = a + 1
            elif "Foundation" in string:
                list_gwp_ftg.append(float(df4['Sum of Global Warming Potential Total (kgCO2eq)'].values[a]))
                a = a + 1
            elif "Footing" in string:
                list_gwp_ftg.append(float(df4['Sum of Global Warming Potential Total (kgCO2eq)'].values[a]))
                a = a + 1
            elif "offset" in string:
                list_gwp_ftg.append(float(df4['Sum of Global Warming Potential Total (kgCO2eq)'].values[a]))
                a = a + 1
            elif "Center" in string:
                list_gwp_ftg.append(float(df4['Sum of Global Warming Potential Total (kgCO2eq)'].values[a]))
                a = a + 1

            else:
                # # need to add for steel wood etc, for now pass to misc
                # # list_gwp_misc.append(float(sheet10[gwp_value].value))
                a = a + 1
                pass
        else:
            a = a + 1


    total_gwp_col = round((sum(list_gwp_col)) / total_slab_area_m2, 1)
    total_gwp_wall = round((sum(list_gwp_wall)) / total_slab_area_m2, 1)
    total_gwp_floor = round((sum(list_gwp_floor)) / total_slab_area_m2, 1)
    total_gwp_ftg = round((sum(list_gwp_ftg)) / total_slab_area_m2, 1)
    total_gwp_transfer = round((sum(list_gwp_transfer)) / total_slab_area_m2, 1)
    total_gwp_misc = round((total_gwp_per_area-total_gwp_col-total_gwp_wall-total_gwp_floor-total_gwp_ftg-total_gwp_transfer), 1)

    #total_gwp_basementwall = round((sum(list_gwp_basementwall)) / total_slab_area_m2, 1)
    #total_gwp_shearwall = round((sum(list_gwp_shearwall)) / total_slab_area_m2, 1)


    x_value = (total_gwp_col, total_gwp_wall, total_gwp_transfer, total_gwp_floor, total_gwp_ftg, total_gwp_misc)

    labels = ["COLUMN", "WALL", "SLAB", "TRANSFER", "FTG", "MISC"]
    mycolors = ['#3c78d8', '#6aa84f', '#f1c232', '#ff8040', '#999999', '#FF0000']
    explodes = (0, 0, 0.1, 0, 0, 0)

    plt.pie([total_gwp_col, total_gwp_wall, total_gwp_floor, total_gwp_transfer,total_gwp_ftg, total_gwp_misc], explode=explodes,
            labels=labels, colors=mycolors, autopct='%.2f %%',)

    plt.show()


    return x_value


def benchmark_color_residential(y_value):
    if type(y_value) is list:
        y_value = sum(y_value)
    else:
        pass
    if y_value < 280:
       mycolor = "#4a9232"
    elif 280 < y_value < 360:
        mycolor = "#8acf30"
    elif 360 < y_value < 440:
        mycolor = "#b9ec5b"
    elif 440 < y_value < 520:
        mycolor = "#ecf10e"
    elif 520 < y_value < 600:
        mycolor = "#ffae88"
    elif 600 < y_value < 680:
        mycolor = "#ff8040"
    else:
        mycolor = "red"

    return mycolor

def scatter_plot():
    global buildingstage
    X1 = [total_slab_area_m2]
    Y_temp = [total_gwp_per_area]

    listx = []
    r1 = 0
    for cell in sheeta['C']:
        r1 = r1 + 1
        if r1 == 1:
            pass
        else:
            listx.append(cell.value)

    # GWP TO LIST
    listy = []
    t1 = 0
    for string in sheeta['J']:
        t1 = t1 + 1
        if t1 == 1:
            pass
        else:
            listy.append(string.value)

    listx.pop(0)
    listy.pop(0)

    plt.ylabel("EMBODIED CARBON INTENSITY (KgCO2/m^2)", font='Arial')
    plt.xlabel("BUILDING AREA (m^2)", font='Arial')

    # print(listx)
    # print(listy)

    orig_map = plt.cm.get_cmap('viridis')
    reversed_map = orig_map.reversed()

    X = [0, 0]
    Y = [200, 700]

    plt.scatter(X, Y, c=Y, s=0.1, cmap=reversed_map)

    # plt.scatter(listx, listy, c=listy, alpha=0.3, cmap=reversed_map)
    plt.scatter(listx, listy, alpha=0.3, color='black')

    SD_DD = name1.get()
    if SD_DD == "":
        if var1.get() == 1:
            SD_y = Y_temp
            plt.scatter(X1, SD_y, marker="^", s=200, color=benchmark_color_residential(SD_y))
            buildingstage = "SD/DD"
        else:
            pass
    elif SD_DD.isnumeric() is False:
        try:
            if float(SD_DD):
                SD_y = float(SD_DD)
                plt.scatter(X1, SD_y, marker="^", s=50, color=benchmark_color_residential(SD_y))


        except ValueError:
            tkinter.messagebox.showinfo("error", "data entered is not numeric")
    else:
        SD_y = float(SD_DD)
        plt.scatter(X1, SD_y, marker="^", s=50, color=benchmark_color_residential(SD_y))


    BP = name2.get()
    if BP == "":
        if var2.get() == 1:
            BP_y = Y_temp
            plt.scatter(X1, BP_y, marker="D", s=200, color=benchmark_color_residential(BP_y))
            buildingstage = "BP"
        else:
            pass
    elif BP.isnumeric() is False:
        try:
            if float(BP):
                BP_y = float(BP)
                plt.scatter(X1, BP_y, marker="D", s=50, color=benchmark_color_residential(BP_y))


        except ValueError:
            tkinter.messagebox.showinfo("error", "data entered is not numeric")
    else:
        BP_y = float(BP)
        plt.scatter(X1, BP_y, marker="D", s=50, color=benchmark_color_residential(BP_y))


    TENDER = name3.get()
    if TENDER == "":
        if var3.get() == 1:
            TENDER_y = Y_temp
            plt.scatter(X1, TENDER_y, marker="s", s=200, color=benchmark_color_residential(TENDER_y))
            buildingstage = "TENDER"
        else:
            pass
    elif TENDER.isnumeric() is False:
        try:
            if float(TENDER):
                TENDER_y = float(TENDER)
                plt.scatter(X1, TENDER_y, marker="s", s=50, color=benchmark_color_residential(TENDER_y))

        except ValueError:
            tkinter.messagebox.showinfo("error", "data entered is not numeric")
    else:
        TENDER_y = float(TENDER)
        plt.scatter(X1, TENDER_y, marker="s", s=50, color=benchmark_color_residential(TENDER_y))

    IFC = name4.get()
    if IFC == "":
        if var4.get() == 1:
            IFC_y = Y_temp
            plt.scatter(X1, IFC_y, marker="X", s=200, color=benchmark_color_residential(IFC_y))
            buildingstage = "IFC"
        else:
            pass
    elif IFC.isnumeric() is False:
        try:
            if float(IFC):
                IFC_y = float(IFC)
                plt.scatter(X1, IFC_y, marker="X", s=50, color=benchmark_color_residential(IFC_y))

        except ValueError:
            tkinter.messagebox.showinfo("error", "data entered is not numeric")
    else:
        IFC_y = float(IFC)
        plt.scatter(X1, IFC_y, marker="X", s=50, color=benchmark_color_residential(IFC_y))

    plt.grid()
    plt.show()
    win.destroy()



def button_of_dotchar():
    global name1, name2, name3, name4, var1, var2, var3, var4, sheeta, win
    path1 = r"U:\Steve\LCA\data collection\GWP data - Residentail.xlsx"
    wb1 = load_workbook(path1)
    sheeta = wb1["Sheet1"]
    # message box of inputing project status
    win = tk.Toplevel()

    # textbox entry widget
    ttk.Label(win, text="GWP (kgCO2eq/m2)").place(x=190, y=5)

    ttk.Label(win, text="Design Development").place(x=50, y=30)
    name1 = tk.StringVar()
    ttk.Entry(win, width=12, textvariable=name1).place(x=200, y=30)
    var1 = tk.IntVar()
    ttk.Checkbutton(win, variable=var1, onvalue=1, offvalue=0).place(x=15, y=30)

    ttk.Label(win, text="Building Permit").place(x=50, y=70)
    name2 = tk.StringVar()
    ttk.Entry(win, width=12, textvariable=name2).place(x=200, y=70)
    var2 = tk.IntVar()
    ttk.Checkbutton(win, variable=var2, onvalue=1, offvalue=0).place(x=15, y=70)

    ttk.Label(win, text="Tender").place(x=50, y=110)
    name3 = tk.StringVar()
    ttk.Entry(win, width=12, textvariable=name3).place(x=200, y=110)
    var3 = tk.IntVar()
    ttk.Checkbutton(win, variable=var3, onvalue=1, offvalue=0).place(x=15, y=110)

    ttk.Label(win, text="IFC").place(x=50, y=150)
    name4 = tk.StringVar()
    ttk.Entry(win, width=12, textvariable=name4).place(x=200, y=150)
    var4 = tk.IntVar()
    ttk.Checkbutton(win, variable=var4, onvalue=1, offvalue=0).place(x=15, y=150)

    # add button
    ttk.Button(win, text='plot', command=scatter_plot).place(x=200, y=190)


    win.geometry('400x250')
    win.title("Project Status")
    win.mainloop()





def button_of_export_data():
    lista = []
    #print(text_buildingtype.get())
    v = text_buildingtype.get()
    #["Residentail", "Education", "Health Care", "Industrial", "Office", "Other"]
    if v == "Residentail":
        p = r'U:\Steve\LCA\data collection\GWP data - Residentail.xlsx'
    elif v == "Education":
        p = r'U:\Steve\LCA\data collection\GWP data - Education.xlsx'
    elif v == "Health Care":
        p = r'U:\Steve\LCA\data collection\GWP data - Health Care.xlsx'
    elif v == "Industrial":
        p = r'U:\Steve\LCA\data collection\GWP data - Industrial.xlsx'
    elif v == "Office":
        p = r'U:\Steve\LCA\data collection\GWP data - Office.xlsx'
    else:
        p = r'U:\Steve\LCA\data collection\GWP data - Other.xlsx'
    wb_excel = load_workbook(p)
    ws_excel = wb_excel['Sheet1']
    jobnumber = text_jobnumber.get()
    #buildingstage = text_buildingtype.get()
    lista.append(str(jobnumber))
    lista.append(buildingstage)
    lista.append(total_slab_area_m2)
    lista.append(total_gwp_floor)
    lista.append(total_gwp_transfer)
    lista.append(total_gwp_wall)
    lista.append(total_gwp_col)
    lista.append(total_gwp_ftg)
    lista.append(total_gwp_misc)
    lista.append(total_gwp_per_area)
    #print(lista)
    #print(jobnumber, buildingstage, total_slab_area_m2, total_gwp_floor, total_gwp_transfer, total_gwp_wall, total_gwp_col, total_gwp_misc,total_gwp_per_area)
    ws_excel.append(lista)
    wb_excel.save(p)

    # delete csv
    csv_files = glob.glob(os.path.join(temp_csv_path, '*.csv'))
    for file in csv_files:
        os.remove(file)

    tkinter.messagebox.showinfo("showinfo", "data exported")

    app.destroy()




app = customtkinter.CTk()
app.geometry("580x500")
app.title("Glotman Simpson.py")



title_frame = customtkinter.CTkFrame(master=app)
title_frame.place(relx=0.28, rely=0.13, anchor=tkinter.CENTER)

title_label = customtkinter.CTkLabel(master=title_frame, justify=tkinter.LEFT,text="One Pager", width=180, height=60, text_font=("Roboto Medium", -28))
title_label.pack(pady=10, padx=10)

frame_1 = customtkinter.CTkFrame(master=app, width=200, height=300)
frame_1.place(relx=0.28, rely=0.57, anchor=tkinter.CENTER)

logo_frame = customtkinter.CTkFrame(master=app, width=200, height=200)
logo_frame.place(relx=0.7, rely=0.25, anchor=tkinter.CENTER)

logo = ImageTk.PhotoImage(Image.open(r"U:\Steve\LCA\one pager program\03\Glotman-logo.png"))

logo_label = customtkinter.CTkLabel(master=logo_frame, image=logo)
logo_label.pack(pady=10, padx=10)


frame_2 = customtkinter.CTkFrame(master=app, width=200, height=185)
frame_2.place(relx=0.7, rely=0.69, anchor=tkinter.CENTER)

# label = customtkinter.CTkLabel(master=app, text="One Pager", width=180, height=60, text_font=("Roboto Medium", -28))
# label.place(relx=0.28, rely=0.15, anchor=tkinter.CENTER)

#
button_readdata = customtkinter.CTkButton(master=frame_1, text="Load Excel", command=read_excel, width=160, height=60,text_font=("Roboto Medium", -20) )
button_readdata.place(relx=0.50, rely=0.2, anchor=tkinter.CENTER)
#
button_piechart = customtkinter.CTkButton(master=frame_1, text="Create Pie Chart", command=button_of_piechar, width=160, height=60,text_font=("Roboto Medium", -20) )
button_piechart.place(relx=0.50, rely=0.53, anchor=tkinter.CENTER)
#
button_dotchart = customtkinter.CTkButton(master=frame_1, text="Create Scatter Plot", command=button_of_dotchar, width=160, height=60,text_font=("Roboto Medium", -20) )
button_dotchart.place(relx=0.50, rely=0.85, anchor=tkinter.CENTER)
#
#button_exportdata = customtkinter.CTkButton(master=frame_1, text="Export Data", command=button_of_dotchar, width=160, height=60,text_font=("Roboto Medium", -20) )
#button_exportdata.place(relx=0.50, rely=0.80, anchor=tkinter.CENTER)

#
label_jobnumber = customtkinter.CTkLabel(master=frame_2, text="Job number", text_font=("Roboto Medium", -10) )
label_jobnumber.place(relx=0.82, rely=0.15, anchor=tkinter.CENTER)

text_jobnumber = customtkinter.CTkEntry(master=frame_2, width=120, height=30,text_font=("Roboto Medium", -10),)
text_jobnumber.place(relx=0.35, rely=0.15, anchor=tkinter.CENTER)

label_buildingtype = customtkinter.CTkLabel(master=frame_2, text="building type", text_font=("Roboto Medium", -10) )
label_buildingtype.place(relx=0.82, rely=0.4, anchor=tkinter.CENTER)

buildingtype_selection = ["", "Residentail", "Education", "Health Care", "Industrial", "Office", "Other"]
text_buildingtype = customtkinter.CTkComboBox(values=buildingtype_selection, master=frame_2, width=120, height=30,text_font=("Roboto Medium", -10),)
text_buildingtype.place(relx=0.35, rely=0.4, anchor=tkinter.CENTER)

#
#button_buildingtype = customtkinter.CTkButton(master=frame_2, text="Building Type", command=button_of_dotchar, width=160, height=40,text_font=("Roboto Medium", -15) )
#button_buildingtype.place(relx=0.5, rely=0.7, anchor=tkinter.CENTER)


button_exportdata = customtkinter.CTkButton(master=frame_2, text="Export Data", command=button_of_export_data, width=160, height=60,text_font=("Roboto Medium", -20) )
button_exportdata.place(relx=0.50, rely=0.74, anchor=tkinter.CENTER)

app.mainloop()
